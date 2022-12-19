from collections import Counter
from dataclasses import fields
import datetime
from msilib.schema import Component
from typing import cast
import re
from jira import JIRA
from jira.client import ResultList
from jira.resources import Issue
import json,os
from io import StringIO
import pandas as pd
from urllib import response
from urllib.request import urlopen
import requests,json
import  urllib3 as ur
from requests.auth import HTTPBasicAuth
import pandas as pd
from Tools import helper
import openpyxl


# =======================================================================================
# ======================================CONFIG===========================================
# =======================================================================================

headers = {
  'Content-Type': 'application/json',
  'Authorization': 'Bearer <your zephyr api key goes here>'
}


Root = os.path.dirname(os.path.abspath(__file__))

path="{}\\TestData\\".format(Root)

jiraOptions = {'server': "https://example.atlassian.net/"} #put your url here
data={"API_token":"vhjjkjkjklkk"}
jira = JIRA(options=jiraOptions,basic_auth=('<your email goes here>', data['API_token']))

# =======================================================================================
# ====================================JSON READER FOR JIRA TEST==========================
# =======================================================================================

def JsonReaderJiraTest(filename,Environment,Component,Testcases,name):
    with open('{}\\TestData\\{}'.format(path,filename))as f:
        # returns JSON object as
        # a dictionary
        data = json.load(f)
        try:
            # Iterating through the json
            # list
            for First in data:
                if Environment in First:
                    for Second in First[Environment]:
                        if Component in Second:
                            for Third in First[Environment][Component]:
                                if Testcases in Third:
                                    for Fourth in First[Environment][Component][Testcases]:
                                        if name in Fourth:

                                            return (First[Environment][Component][Testcases][name])
        except:
            return "Not Found"
# =======================================================================================
# ====================================JSON WRITER FOR JIRA TEST==========================
# =======================================================================================

def JiraDataToJson(Environment,raw,SubEnv,SubsubEnv):


        with open("{}\\Jira.json".format(path)) as file:
            load = json.load(file)
        file.close()
    
        try:
            for Env in load:
                if Environment in Env:
                    for EnvEnv in Env:
                        if SubEnv in EnvEnv:
                            for SubsubEnv in EnvEnv:
                                if SubEnv in SubsubEnv:
                                    Env[Environment]={SubEnv:{SubsubEnv:raw  }   }  
                                else: 
                                        Env[Environment][SubEnv]={SubsubEnv:raw  } 
                        else:
                             Env[Environment][SubEnv][SubsubEnv]=raw 
                                    
                else:
                     
                    Env[Environment][SubEnv][SubsubEnv]=raw   
        except:
               for Env in load:
                if Environment in Env:
                    for EnvEnv in Env:
                        if SubEnv in EnvEnv:
                            for SubsubEnv in EnvEnv:
                                if SubEnv in SubsubEnv:
                                    Env[Environment]={SubEnv:{SubsubEnv:raw  }   }  
                                else: 
                                        Env[Environment][SubEnv]={SubsubEnv:raw  } 
                        else:
                             Env[Environment][SubEnv]={SubsubEnv:raw  }  
                                    
                else:
                     
                    Env[Environment]={SubEnv:{SubsubEnv:raw  }   }  
         
            #print(data)
        with open("{}Jira.json".format(path), "w+") as file:
            json.dump(load,file,indent=4)
        file.close()

# =======================================================================================
# ======================================GET ZEPHYR FOLDERS===============================
# =======================================================================================

def GetFolders():
    request = requests.get('https://api.zephyrscale.smartbear.com/v2/folders',headers=headers )
    return request.content

# =======================================================================================
# =================================CREATE ISSUE AND EXECUTION LINK=======================
# =======================================================================================

def CreateissuelinkExecution(testExecutionIdOrKey,issueId):
    """Create a new test case Cycle."""
    data={
  "issueId": issueId
}

    response=requests.post(f"https://api.zephyrscale.smartbear.com/v2/testexecutions/{testExecutionIdOrKey}/links/issues",headers=headers,data=json.dumps(data))
    return response.json()

# =======================================================================================
# ==================================CREATE ISSUE AND TESTCASE LINK=======================
# =======================================================================================

def Createissuelink(testCaseKey,issueId):
    """Create a new test case Cycle."""
    data={
  "issueId": issueId
}

    response=requests.post(f"https://api.zephyrscale.smartbear.com/v2/testcases/{testCaseKey}/links/issues",headers=headers,data=json.dumps(data))
    return response.json()

# =======================================================================================
# ======================================CREATE TEST SCRIPTS==============================
# =======================================================================================

def CreateScript(testCaseKey,Steps,description,TestData,ExpectedResult):
    row={}
    row["Steps"]=Steps
    row["ExpectedResult"]=ExpectedResult
    data={
"type": "plain",
"text":json.dumps(row,indent=4),


}
    
    response=requests.post(f"https://api.zephyrscale.smartbear.com/v2/testcases/{testCaseKey}/testscript",headers=headers,data=json.dumps(data))
    print(response.content)
    return response.json()['id']


# =======================================================================================
# ===================================CREATE ISSUE AND CYCLE LINK=========================
# =======================================================================================

def IssuelinktestCycle(testCycleId,issueId):
  data={
"issueId": issueId
}
  response=requests.post(f"https://api.zephyrscale.smartbear.com/v2/testcycles/{testCycleId}/links/issues",headers=headers,data=json.dumps(data))
  print(response.content)

# =======================================================================================
# =====================CREATE FINAL TEST EXECUTION FROM EXCEL SHEETS=====================
# =======================================================================================

def CreateFinalTestExecution(fileName,Sheet_Name):
    print("Executing TestCases ...")
    excel_data_df = pd.read_excel('{}{}'.format(path,fileName), sheet_name=Sheet_Name)

    json_str = excel_data_df.to_json(orient='records')
    thisisjson_dict = json.loads(json_str)
    for row in thisisjson_dict:

        TestIssueKey = JsonReaderJiraTest("Jira.json", Sheet_Name, "Issues", row["Test Case ID"],
                                                      "Key")
        TestCyclesKey = JsonReaderJiraTest("Jira.json", Sheet_Name, "TestCycles", Sheet_Name,
                                          "Key")
        TestIssueId = JsonReaderJiraTest("Jira.json", Sheet_Name, "Issues", row["Test Case ID"],
                                                     "Id")
        TestcaseKey = JsonReaderJiraTest("Jira.json", Sheet_Name, "Testcases", row["Test Case ID"], "Key")
        TestCycleKey = JsonReaderJiraTest("Jira.json", Sheet_Name, "TestCycles", Sheet_Name, "Key")
        try:
            CreateScript(TestcaseKey,row["Test Steps"],row["Test Case Description"],row["Expected Results"])
        except:
            pass
        data={
            "projectKey": "BT",
            "testCaseKey": TestcaseKey,
            "testCycleKey": TestCycleKey,
            "statusName": row["Pass/Skip/Fail"],
            "testScriptResults": [
            {
            "statusName": row["Pass/Skip/Fail"],

            "actualResult": row["Actual Results"]
            }
            ],

            "executionTime": 120000,
            "executedById": "user.id",
            "assignedToId": "user.id",
            "comment": row["Actual Results"]

        }
        response = requests.post("https://api.zephyrscale.smartbear.com/v2/testexecutions", headers=headers,
                                 data=json.dumps(data))
        transition(TestIssueId, "Done", Sheet_Name, fields=None)
        print(response.content)
    #return response.json()["id"]

# =======================================================================================
# ==================================CREATE SINGLE TEST EXECUTION=========================
# =======================================================================================

def CreateTestExecution(IssueID,testCaseKey,testCycleKey,statusName,TestData,actualResult):
    print(testCaseKey)
    data={
"projectKey": "BT",
"testCaseKey": str(testCaseKey),
"testCycleKey": testCycleKey,
"statusName": statusName,
"testScriptResults": [
{
"statusName": statusName,

"actualResult": actualResult
}
],

"executionTime": 120000,
"executedById": "user.id",
"assignedToId": "user.id",
"comment": actualResult

    } 
    response = requests.post("https://api.zephyrscale.smartbear.com/v2/testexecutions",headers=headers,data=json.dumps(data))
    transition(IssueID, "Done", "Tested", fields=None)
    print( response.content)
    return response.json()["id"]

# =======================================================================================
# ====================================VIEW ALL TEST EXECUTIONS===========================
# =======================================================================================

def GetAllTestExecutions():
    response=requests.get("https://api.zephyrscale.smartbear.com/v2/testexecutions",headers=headers)
    print(response.content)

# =======================================================================================
# ======================================CREATE TEST CYCLE================================
# =======================================================================================

def CreateTestCycle(Sheet_Name):
    raw={}
  
    """Create a test case cycle."""
    data={
    "projectKey": "BT",
    "name": Sheet_Name,
    "description": "Rider Login",
    

    "statusId": 2606208,
    
    "ownerId": "user.id",

    }

    #"plannedStartDate": str(datetime.datetime.now().strftime("%A %d. %B %Y, %H:%M:%S")),
    #"plannedEndDate": str(datetime.datetime.now().strftime("%A %d. %B %Y, %H:%M:%S")),
    response=requests.post(f"https://api.zephyrscale.smartbear.com/v2/testcycles",headers=headers,data=json.dumps(data))
    raw["Id"] = response.json()['id']
    raw["Key"] = response.json()['key']
    JiraDataToJson(Sheet_Name,raw,"TestCycles",Sheet_Name)
    print(response.content)
    return response.json()['key']

# =======================================================================================
# ====================================VIEW ALL TEST CYCLE================================
# =======================================================================================

def GetAllLifeCycle():
    LifeCycle=[]
    response = requests.get("https://api.zephyrscale.smartbear.com/v2/testcycles",headers=headers)
    print(response.json()["values"])
    for name in response.json()["values"]:
        LifeCycle.append(name["name"])
    return LifeCycle

# =======================================================================================
# ================================CHECK EXISTANCE OF LIFE CYCLE==========================
# =======================================================================================

def Existance_of_LifeCycle(Sheet_Name):
    response = requests.get(f"https://api.zephyrscale.smartbear.com/v2/testcycles",headers=headers)
    for name in response.json()["values"]:
        print(response.json())
        if Sheet_Name in name['name']:
            return name['id']
    return 0

# =======================================================================================
# ====================================GET ALL CYCLE==========================
# =======================================================================================

def GetCycle(Name):
    response = requests.get(f"https://api.zephyrscale.smartbear.com/v2/testcycles",headers=headers)
    return (response.json())

# =======================================================================================
# =====================================CREATE TESTCASE==================================
# =======================================================================================

def CreateTestCase(Sheet_Name,IssueID,TestCycleId,TestCaseID,TestCaseDescription,TestSteps,TestData,ExpectedResult,ActualResult,TestStatus):
    raw={}
    folderId=""
    if "Rider" in Sheet_Name:
        folderId="3346181"
    elif "Customer" in Sheet_Name:
        folderId = "3498063"
    elif "Restaurant" in Sheet_Name:
        folderId = "5128064"
    else:
        folderId = "3292689"

    data3={
      "projectKey": "BT",
      "name": TestCaseID,
      "objective": TestCaseDescription,
      "precondition": TestData,
      "estimatedTime": 138000,
      "componentId": 10001,
      "priorityId":2606212,
      "statusId": 2606208,
      "folderId":folderId,
      "labels": [
      "Regression",
      "Performance",
      "Automated"
      ]

      }
    response=requests.post('https://api.zephyrscale.smartbear.com/v2/testcases',headers=headers,data=json.dumps(data3))
    jsonres=response.json()
    key=jsonres["key"]
    id=jsonres["id"]
    raw["Id"]=id
    raw["Key"]=key
    raw["Name"]=TestCaseID
    raw["IssueId"]=IssueID
    JiraDataToJson(Sheet_Name,raw,"Testcases",TestCaseID)
    Createissuelink(key,IssueID)
    CreateScript(key,TestSteps,TestCaseDescription,TestData,ExpectedResult)
    IssuelinktestCycle(TestCycleId,IssueID)
    CreateTestExecution(IssueID,key,TestCycleId,TestStatus,TestData,ActualResult)

# =======================================================================================
# ====================================CREATE ISSUE==========================
# =======================================================================================

def CreateIssue(Sheet_Name,description):
    payload=json.dumps({
        "fields": {
        "project":
        {
            "key": "BT"
        },
        "summary": Sheet_Name,
        "description": description,
        "issuetype": {
            "name": "Task"
        }
    }
    })
    new_issue = jira.create_issue(fields=payload)

# =======================================================================================
# ====================================DELETE ISSUE==========================
# =======================================================================================

def DeleteIssue(IssueId):
    issue=jira.issue(IssueId)
    issue.delete()

def ExistanceofIssue(IssueId):
    try:
        issue = jira.issue(IssueId)
        if int(issue.id)>0:
            return issue.fields.summary
    except:
        return 0

# =======================================================================================
# =======================================FIND ISSUE======================================
# =======================================================================================

def FindIssues(IssueId):
    try:
        issue = jira.issue(IssueId)
        if int(issue.id)>0:
            return issue
    except:
        return "Not Found"

# =======================================================================================
# ======================================UPDATE ISSUE=====================================
# =======================================================================================

def UpdateIssue(IssueId,description,Summmary,notify):
    try:
        issue = jira.issue(IssueId)
        issue.update(notify=False,
            summary=Summmary, description=description)
        return issue.id
    except:
        return None

# =======================================================================================
# ====================================ADD COMMENT========================================
# =======================================================================================

def IssueAddComment(IssueId,Comment):
    jira.add_comment(IssueId, Comment)


def ExistanceOfIssuesByName(SheetName):
    initial = 0
    size = 100
    start = initial * size
    issues = jira.search_issues('project=10000', start, size)

    initial += 1

    for issue in issues:

        try:
            if SheetName in issue.fields.summary:
                return issue.fields.summary
        except:
            pass
    return 0
# =======================================================================================
# ======================================gET ALL ISSUES===================================
# =======================================================================================

def GetIssues():
    initial = 0
    size = 100
    start= initial*size
    issues = jira.search_issues('project=10000',  start,size)


    initial += 1

    for issue in issues:
        
        try:

            #print(issue.raw)

           # if(issue.fields.parent.id=="10033"):
                print("parent-id=", issue.fields.parent.id)
                print("Issue-no=", issue.id)
                print('IssueType=', issue.fields.issuetype.name)
                print('Status=', issue.fields.status.name)
                print('Summary=', issue.fields.summary)


        
                print ("================================================")
                print ("================================================")
                print ("================================================")
        except:
            pass
            # print("Issue-Id=", issue.id)
            # print('IssueType=', issue.fields.issuetype.name)
            # print('Status=', issue.fields.status.name)
            # print('Summary=', issue.fields.summary)
            #
            # print("================================================")
            # print("================================================")
            # print("================================================")
        #finally:
            #print(issues)

# =======================================================================================
# ====================================CREATE FOLDER======================================
# =======================================================================================

def CreateFolder(ParentDir,DirectoryName):

    path = os.path.join(ParentDir, DirectoryName)
    try:
        os.mkdir(path)
        print("Directory '% s' created" % DirectoryName)
    except Exception as ex:
        print("Failed {}".format(ex))

# =======================================================================================
# ====================DELETE ALL ATTACHMENTS RELATED TO CURRENT ISSUE====================
# =======================================================================================

def DeleteIssuesAttachments(IssueId):
    query=jira.issue(IssueId)
    try:
        for a in query.fields.attachment:
                print("For issue {}, found attach: '{}' [{}].".format(query.key, a.filename, a.id))
                print(a.id)
                jira.delete_attachment(a.id)
    except:
        pass

# =======================================================================================
# ============================ATTACH FILES TO THE CURRENT ISSUE==========================
# =======================================================================================

def IssuesAttachfiles(IssueId,ImgList,Path):
    DeleteIssuesAttachments(IssueId)
    if len(ImgList)>0:
        for image in ImgList:
            jira.add_attachment(issue=IssueId,
                                attachment='{}{}'.format(Path,image))

# =======================================================================================
# ====================================CREATE COMPONENT===================================
# =======================================================================================

def component():
    bulk_proj = jira.project('BT')
    component = jira.create_component('Test Component', bulk_proj, description='testing!!', leadUserName='Abinet',
            assigneeType='PROJECT_LEAD', isAssigneeTypeValid=False)

# =======================================================================================
# ====================================CREATE ISSUE ALTERNATE ==========================
# =======================================================================================

def create_issue_with_fielddict(summary,description):
        
        fields = {
            'project': {
                'key': 'BT'
            },
            'summary':summary ,
            'description': description,
            'issuetype': {
                'name': 'Task'
            }
           
        }
        issue = jira.create_issue(fields=fields)
        return issue.id

# =======================================================================================
# ====================================CREATE EPIC CHILD==========================
# =======================================================================================


def Create_Epic_Child(parentid,customfield_10040,Sheet_Name,Description,IssueType):
    raw={}

    subtask_one = {
        'project': {'key': 'BT'},
        "summary": Sheet_Name,
        "description": Description,
        'issuetype': {'name': IssueType},
        'parent': {'id': parentid},
        'assignee': {'name': 'user.name'},
        "customfield_10040":Sheet_Name,

    }
    # print(singleIssue.id)

    child = jira.create_issue(fields=subtask_one)
    return child.id

# =======================================================================================
# =================================CREATE ISSUE FROM EXCEL DATA==========================
# =======================================================================================

def create_Child_issue_with_fielddict(filename,parentid,Sheet_Name):
    description=""
    try:
        raw={}
        excel_data_df = pd.read_excel('{}{}.xlsx'.format(path,filename), sheet_name=Sheet_Name)
        TestCycleId=CreateTestCycle(Sheet_Name)
        json_str = excel_data_df.to_json(orient='records')
        thisisjson_dict =json.loads(json_str)
        for row in thisisjson_dict:
            description=row["Test Case Description"]
            subtask_one = {
                'project' : { 'key' : 'BT' },
                "summary": row["Test Case ID"],
                "description": row["Test Case Description"],
                'issuetype' : { 'name' : 'Subtask' },
                'parent' : { 'id' : parentid },
                'assignee' : { 'name' : 'user.name'},
            }
        #print(singleIssue.id)
            if len(jira.issue(parentid).id) > 0 :
                    child = jira.create_issue(fields=subtask_one)
                    raw["Id"]=child.id
                    raw["Key"]=child.key
                    raw["ParentId"]=parentid
                    raw["Name"]=row["Test Case ID"]
                    JiraDataToJson(Sheet_Name,raw,"Issues",row["Test Case ID"])
                    CreateTestCase(Sheet_Name,child.id,TestCycleId,row["Test Case ID"],row["Test Case Description"],row["Test Steps"],row["Test Data"],row["Expected Results"],row["Actual Results"],row["Pass/Skip/Fail"])

                    print("created child with id : " + child.id)


            elif len(jira.issue(parentid).id) == 0:
                print ("the input was empty")
            else:
                print ("The following input was not valid" + jira.issue(parentid).id)
                #In Progress
        transition(parentid, "Done", "Tested", fields=None)
    except:
        raw = {}
        excel_data_df = pd.read_excel('{}{}.xlsx'.format(path,filename), sheet_name=Sheet_Name)
        TestCycleId = CreateTestCycle(Sheet_Name)
        json_str = excel_data_df.to_json(orient='records')
        thisisjson_dict = json.loads(json_str)
        for row in thisisjson_dict:
            description = row["Test Case Description"]
            subtask_one = {
                'project': {'key': 'BT'},
                "summary": row["Test Case ID"],
                "description": row["Test Case Description"],
                'issuetype': {'name': 'Subtask'},
                'parent': {'id': parentid},
                'assignee': {'name': 'user.name'},
            }
            # print(singleIssue.id)
            if len(jira.issue(parentid).id) > 0 and len(jira.issue(parentid).id) <= 7:
                child = jira.create_issue(fields=subtask_one)
                raw["Id"] = child.id
                raw["Key"] = child.key
                raw["ParentId"] = parentid
                raw["Name"] = row["Test Case ID"]
                JiraDataToJson(Sheet_Name, raw, "Issues", row["Test Case ID"])
                CreateTestCase(Sheet_Name, child.id, TestCycleId, row["Test Case ID"], row["Test Case Description"],
                               row["Test Steps"], row["Test Data"], row["Expected Results"], row["Actual Results"],
                               row["Pass/Skip/Fail"])

                print("created child with id : " + child.id)


            elif len(jira.issue(parentid).id) == 0:
                print("the input was empty")
            else:
                print("The following input was not valid" + jira.issue(parentid).id)
                # In Progress
        transition(parentid, "Done", "Tested", fields=None)


# =======================================================================================
# ====================================GET TRANSITION BY NAME ==========================
# =======================================================================================

def _get_tr_id_by_name(trans_available, name):
        '''Get transition id by target state name'''
        for trans in trans_available:
            if trans['to']['name'].lower() == name.lower():
                return trans['id']
        return None

# =======================================================================================
# =======================================TRANSITION======================================
# =======================================================================================

def transition(issue, state, comment, fields=None):
    '''Transition field with comment to specific state'''
    trans_available = jira.transitions(issue)
    trans_id = _get_tr_id_by_name(trans_available, state)
    _issue = jira.issue(issue)
    if trans_id:
        jira.transition_issue(_issue, trans_id, comment=comment, fields=fields)
        return True
    else:
        return False
# =======================================================================================
# =============================CREATE TESTCASE FROM EXCEL SHEET==========================
# =======================================================================================

def CreateTestCasesFromExcel(fileName,ParentId):
    wb = openpyxl.load_workbook('{}{}.xlsx'.format(path,fileName))
    for sheetName in wb.sheetnames:
        #print(sheetName)

            if ExistanceOfIssuesByName(sheetName)==0:
                Child = Create_Epic_Child(ParentId, sheetName, sheetName, sheetName, "Task")
                print("New Epic Child Created With Id: {}".format(Child))
                create_Child_issue_with_fielddict(fileName, Child, sheetName)
            else:
                print("==========================")
                print("==========================")
                print("==========================")
                print("{} Exist".format(sheetName))
                print("==========================")
                print("==========================")
                print("==========================")

# =======================================================================================
# ========================CREATE RIDER FOLDERS FROM EXCEL SHEET==========================
# =======================================================================================

def CreateRiderFolder():
    wb = openpyxl.load_workbook('{}TestcasesRider.xlsx'.format(path))
    for sheetName in wb.sheetnames:
        try:
            CreateFolder("{}{}".format(Root,"Exported\\Mobile\\Rider\\images\\"),sheetName)
        except:
            pass

# =======================================================================================
# =======================CREATE CUSTOMER FOLDERS FROM EXCEL SHEET========================
#========================================================================================

def CreateCustomerFolder():
    wb = openpyxl.load_workbook('{}TestcasesCustomer.xlsx'.format(path))
    for sheetName in wb.sheetnames:
        try:
            CreateFolder("{}{}".format(Root,"Exported\\Mobile\\Customer\\images\\"),sheetName)
        except:
            pass

# =======================================================================================
# ========================CREATE ADMIN FOLDERS FROM EXCEL SHEET==========================
# =======================================================================================


def CreateAdminFolder():
    wb = openpyxl.load_workbook('{}TestcasesAdmin.xlsx'.format(path))
    for sheetName in wb.sheetnames:
        try:
            CreateFolder("{}{}".format(Root,"Exported\\Admin\\images\\"),sheetName)
        except:
            pass


# =======================================================================================
# ========================CREATE Restaurant FOLDERS FROM EXCEL SHEET==========================
# =======================================================================================


def RestaurantFolder():
    wb = openpyxl.load_workbook('{}TestcasesRestaurant.xlsx'.format(path))
    for sheetName in wb.sheetnames:
        try:
            CreateFolder("{}{}".format(Root,"Exported\\Mobile\\Restaurant\\images\\"),sheetName)
        except:
            pass

# =======================================================================================
# =======================================================================================
#CreateCustomerFolder()
#RestaurantFolder()
#CreateAdminFolder()
#CreateRiderFolder()
#RestaurantFolder()
#"TestcasesRestaurant","10471"
#"TestcasesCustomer","10336"
#"TestcaseAdmin","10399"
#"TestcasesRider","10033"

#CreateTestCasesFromExcel("TestcasesCustomer","10336")
# wb = openpyxl.load_workbook('{}{}.xlsx'.format(path,"TestcasesCustomer"))
# for sheetName in wb.sheetnames:
#GetAllLifeCycle()
#print(Existance_of_LifeCycle("CustomerApp"))
#print(ExistanceofIssue(10033))
#CreateFinalTestExecution("TestcasesCustomer.xlsx","CustomerApp")
#GetIssues()
#DeleteIssuesAttachments("10378")
#create_Child_issue_with_fielddict(str(10034),Sheet_Name)
#print(GetFolders())
#CreateTestCycle(Sheet_Name)
#GetAllLifeCycle()
#GetAllTestExecutions()
#print(IssuelinktestCycle("BT-R13",10070))

# Test_Data = helper.EXCELREADER("TestcasesRider.xlsx", Sheet_Name, "Test Data")
# Test_Status = helper.EXCELREADER("TestcasesRider.xlsx", Sheet_Name, "Pass/Skip/Fail")
# Test_Case=helper.EXCELREADER("TestcasesRider.xlsx", Sheet_Name, "Test Case ID")
# TestcaseKey=JsonReaderJiraTest("Jira.json",Sheet_Name,"Testcases",Test_Case[0],"Key")
# TestCycleId=JsonReaderJiraTest("Jira.json",Sheet_Name,"TestCycles",Sheet_Name,"Key")
# TestStatus=Test_Status[0]
#
# ActualResult=helper.EXCELREADER("TestcasesRider.xlsx", Sheet_Name, "Actual Results")[0]
# CreateTestExecution(TestcaseKey,TestCycleId,TestStatus,Test_Data[0],ActualResult)
